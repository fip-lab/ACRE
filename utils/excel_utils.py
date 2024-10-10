#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

"""
@File    : file_utils.py
@Author  : huanggj
@Time    : 2023/2/17 9:31
"""
import pandas as pd
import os
import glob
from openpyxl import load_workbook

# 读取Excel任务
def read_excel(file_path, column_type_dict , sheet_name=0):
    # 使用pandas读取excel文件
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=1, engine='openpyxl')
    # 转换数据类型
    for column_name,target_type  in column_type_dict.items():
        try:
            df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
            df = df.dropna(subset=[column_name])
            df[column_name] = df[column_name].astype(target_type)

            #df[column_name] = df[column_name].astype(target_type)
        except Exception as e:
            print(e)
            print(f"Cannot convert column '{column_name}' to {target_type}")
    # 将DataFrame转化为字典列表，每一行是一个任务，列名是参数名，列值是参数值
    tasks = df.to_dict(orient='records')
    # 把 Nan值转换为None
    for i in range(len(tasks)):
        for key, value in tasks[i].items():
            if pd.isnull(value):
                tasks[i][key] = None
    return tasks



# 获取最新的Excel文件
def get_latest_excel(dir_path):
    if not os.path.isdir(dir_path):
        raise Exception(f"{dir_path} 不是一个文件夹")
    # 获取目标目录下所有的Excel文件
    files = glob.glob(os.path.join(dir_path, "*.xlsx")) + glob.glob(os.path.join(dir_path, "*.xls"))
    # 如果目录下没有Excel文件，返回None
    if not files:
        raise  Exception(f"path : {dir_path}  no excel file")
    # 按照修改时间排序，获取修改时间最新的文件
    latest_file = max(files, key=os.path.getmtime)

    if latest_file is not None:
        print(f"Latest Excel file is: {latest_file}")

    return latest_file

# 写入实验任务表结果
def write_result(excel_file, data, column_type_dict):
    # 数据
    task_id = data['task_id']
    start_time = data['start_time']
    end_time = data['end_time']
    res = data['result']

    # 读取Excel文件
    df = pd.read_excel(excel_file, sheet_name=0, header=1, engine='openpyxl')
    # 转换数据类型
    for column_name, target_type in column_type_dict.items():
        try:
            df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
            df = df.dropna(subset=[column_name])
            df[column_name] = df[column_name].astype(target_type)

            # df[column_name] = df[column_name].astype(target_type)
        except Exception as e:
            print(e)
            print(f"Cannot convert column '{column_name}' to {target_type}")

    # 找到task_id对应的行号
    row_num = df.index[df['task_id'] == task_id].tolist()[0] + 3  # 加2是因为pandas的索引从0开始，而Excel的行号从1开始，同时Excel可能还有表头行

    # 加载Workbook
    wb = load_workbook(excel_file)

    # 选择工作表
    ws = wb.worksheets[0]

    # 写入数据
    ws.cell(row=row_num, column=3).value = '√'
    ws.cell(row=row_num, column=4).value = start_time
    ws.cell(row=row_num, column=5).value = end_time
    ws.cell(row=row_num, column=6).value = str(res) + "%"

    # 保存文件
    wb.save(excel_file)
    wb.close()

# 回填大表
def write_main_table(file_path, res_list):
    # 加载Workbook
    wb = load_workbook(file_path)

    # 选择工作表
    ws = wb.worksheets[0]  # 假设总体任务表格在一个名为"task_summary"的工作表中

    # 在工作表中找到task_id对应的单元格并写入值
    for item in res_list:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == item['task_id']:
                    ws.cell(row=cell.row, column=cell.column + 1, value=item['value'])  # 假设你需要写入的列是task_id所在的列的下一列

    # 保存文件
    wb.save(file_path)
    wb.close()